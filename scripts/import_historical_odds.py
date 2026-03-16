#!/usr/bin/env python
"""Import MLB historical odds into games + odds_history.

Primary source:
  - local XLSX workbook (`mlb-odds.xlsx`) using openpyxl

Optional source:
  - normalized CSV exports from free scrapers
    (mlb-odds-scraper / sportsbookreview-scraper adapters)
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from zoneinfo import ZoneInfo

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover - runtime guard
    raise SystemExit("openpyxl is required; install with `python -m pip install openpyxl`") from exc

try:
    import psycopg
except ImportError as exc:  # pragma: no cover - runtime guard
    raise SystemExit("psycopg is required; install with `python -m pip install psycopg[binary]`") from exc


ET_ZONE = ZoneInfo("America/New_York")

DB_SOURCE = "mlb-free-odds"
DEFAULT_BOOK_KEY = "consensus_xlsx"
DEFAULT_BOOK_NAME = "XLSX Consensus"

SOURCE_PRIORITY = {
    "mlb-odds-xlsx": 1,
    "sportsbookreview-scraper": 2,
    "mlb-odds-scraper": 3,
}


@dataclass
class MarketRow:
    market_key: str
    market_name: str
    outcome_name: str
    outcome_side: str
    price_american: int
    point: Optional[float]


@dataclass
class GameOddsRecord:
    external_id: str
    sport: str
    home_team: str
    away_team: str
    commence_time: datetime
    source: str
    upstream_source: str
    priority: int
    book_key: str
    book_name: str
    markets: List[MarketRow]
    raw_payload: Dict[str, Any]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import MLB historical odds into betbot")
    parser.add_argument(
        "--xlsx",
        default=r"C:\Users\Admin\Downloads\mlb-odds.xlsx",
        help="Path to MLB odds workbook",
    )
    parser.add_argument(
        "--skip-xlsx",
        action="store_true",
        help="Skip workbook import and ingest scraped CSV only",
    )
    parser.add_argument(
        "--sheet",
        default="Betting Odds",
        help="Workbook sheet containing odds rows",
    )
    parser.add_argument(
        "--scraped-csv",
        action="append",
        default=[],
        help="Optional normalized scraper CSV path (repeatable)",
    )
    parser.add_argument(
        "--database-url",
        default=os.getenv("BETBOT_DATABASE_URL", ""),
        help="Postgres connection string (defaults to BETBOT_DATABASE_URL)",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate rows without writing to database",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.database_url and not args.dry_run:
        raise SystemExit("database URL required via --database-url or BETBOT_DATABASE_URL")

    records: List[GameOddsRecord] = []
    if not args.skip_xlsx:
        records.extend(load_xlsx_records(Path(args.xlsx), args.sheet))
    for csv_path in args.scraped_csv:
        records.extend(load_normalized_scraped_csv(Path(csv_path)))
    if not records:
        raise SystemExit("no records loaded; provide XLSX input and/or --scraped-csv files")

    # Deterministic ordering keeps replay artifacts stable across reruns.
    records.sort(key=lambda r: (r.priority, r.commence_time, r.away_team, r.home_team, r.book_key))
    print(f"parsed records: {len(records)}")

    if args.dry_run:
        print("dry-run enabled; no writes performed")
        return

    inserted_games = 0
    inserted_snapshots = 0
    skipped_snapshots = 0
    skipped_precedence = 0
    seen_precedence: Dict[Tuple[str, str, str, str, Optional[float]], int] = {}

    with psycopg.connect(args.database_url) as conn:
        with conn.cursor() as cur:
            for record in records:
                game_id, created = upsert_game(cur, record)
                if created:
                    inserted_games += 1
                for market in record.markets:
                    precedence_key = market_precedence_key(record, market)
                    existing_priority = seen_precedence.get(precedence_key)
                    if existing_priority is not None and existing_priority <= record.priority:
                        skipped_precedence += 1
                        continue
                    inserted, skipped = insert_snapshot_if_changed(cur, game_id, record, market)
                    if inserted:
                        inserted_snapshots += 1
                    if skipped:
                        skipped_snapshots += 1
                    seen_precedence[precedence_key] = record.priority
        conn.commit()

    print(f"games_upserted: {inserted_games}")
    print(f"snapshots_inserted: {inserted_snapshots}")
    print(f"snapshots_skipped_dedup: {skipped_snapshots}")
    print(f"snapshots_skipped_precedence: {skipped_precedence}")


def load_xlsx_records(path: Path, sheet_name: str) -> List[GameOddsRecord]:
    if not path.exists():
        raise SystemExit(f"xlsx file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"sheet {sheet_name!r} not found in workbook")

    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    normalized_headers = {name: idx for idx, name in enumerate(headers)}

    required = {
        "Date",
        "Start Time (EDT)",
        "Away",
        "Home",
        "Status",
        "Away ML",
        "Home ML",
    }
    missing = sorted(required - set(normalized_headers))
    if missing:
        raise SystemExit(f"xlsx missing required columns: {', '.join(missing)}")

    out: List[GameOddsRecord] = []
    for idx, row in enumerate(rows, start=2):
        row_values = [row[i] if i < len(row) else None for i in range(len(headers))]
        payload = {name: row_values[pos] for name, pos in normalized_headers.items()}
        try:
            record = record_from_row(payload, source="mlb-odds-xlsx", book_key=DEFAULT_BOOK_KEY, book_name=DEFAULT_BOOK_NAME)
        except ValueError as exc:
            raise SystemExit(f"xlsx parse error row {idx}: {exc}") from exc
        if record is None:
            continue
        out.append(record)
    return out


def load_normalized_scraped_csv(path: Path) -> List[GameOddsRecord]:
    if not path.exists():
        raise SystemExit(f"scraped CSV not found: {path}")

    with path.open("r", newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            raise SystemExit(f"scraped CSV has no headers: {path}")

        out: List[GameOddsRecord] = []
        for idx, row in enumerate(reader, start=2):
            source = (row.get("source") or "scraped-csv").strip().lower()
            book_key = slugify(row.get("book_key") or row.get("book") or "scraped")
            book_name = (row.get("book_name") or row.get("book") or "Scraped Book").strip()
            try:
                record = record_from_row(row, source=source, book_key=book_key, book_name=book_name)
            except ValueError as exc:
                raise SystemExit(f"scraped CSV parse error {path}:{idx}: {exc}") from exc
            if record is None:
                continue
            out.append(record)
    return out


def record_from_row(raw: Dict[str, Any], source: str, book_key: str, book_name: str) -> Optional[GameOddsRecord]:
    status = str(raw.get("Status") or raw.get("status") or "").strip()
    # Keep only rows that have at least moneyline values.
    away_ml = parse_american(raw.get("Away ML") or raw.get("away_ml"))
    home_ml = parse_american(raw.get("Home ML") or raw.get("home_ml"))
    if away_ml is None or home_ml is None:
        return None

    away_team = normalize_team(str(raw.get("Away") or raw.get("away") or "").strip())
    home_team = normalize_team(str(raw.get("Home") or raw.get("home") or "").strip())
    if not away_team or not home_team:
        raise ValueError("away/home team is required")

    game_date = parse_game_date(raw.get("Date") or raw.get("date"))
    start_time = str(raw.get("Start Time (EDT)") or raw.get("start_time_edt") or raw.get("start_time") or "7:05 PM")
    commence_time = parse_commence_time(game_date, start_time)

    external_id = build_external_id(game_date, away_team, home_team)

    markets = [
        MarketRow("h2h", "Moneyline", away_team, "away", away_ml, None),
        MarketRow("h2h", "Moneyline", home_team, "home", home_ml, None),
    ]

    total_line = parse_float(raw.get("O/U") or raw.get("total"))
    over_price = parse_american(raw.get("Over") or raw.get("over"))
    under_price = parse_american(raw.get("Under") or raw.get("under"))
    if total_line is not None and over_price is not None and under_price is not None:
        markets.extend(
            [
                MarketRow("totals", "Total Runs", "Over", "over", over_price, total_line),
                MarketRow("totals", "Total Runs", "Under", "under", under_price, total_line),
            ]
        )

    home_spread = parse_float(raw.get("Home RL Spread") or raw.get("home_rl_spread"))
    rl_away = parse_american(raw.get("RL Away") or raw.get("rl_away"))
    rl_home = parse_american(raw.get("RL Home") or raw.get("rl_home"))
    if home_spread is not None and rl_away is not None and rl_home is not None:
        away_spread = -home_spread
        markets.extend(
            [
                MarketRow("spreads", "Run Line", away_team, "away", rl_away, away_spread),
                MarketRow("spreads", "Run Line", home_team, "home", rl_home, home_spread),
            ]
        )

    payload = dict(raw)
    payload["normalized_status"] = status
    payload["upstream_source"] = source
    payload["book_key"] = book_key
    payload["book_name"] = book_name
    payload["imported_at_utc"] = datetime.now(timezone.utc).isoformat()

    return GameOddsRecord(
        external_id=external_id,
        sport="MLB",
        home_team=home_team,
        away_team=away_team,
        commence_time=commence_time,
        source=DB_SOURCE,
        upstream_source=source,
        priority=source_priority(source),
        book_key=book_key,
        book_name=book_name,
        markets=markets,
        raw_payload=payload,
    )


def upsert_game(cur: psycopg.Cursor, record: GameOddsRecord) -> tuple[int, bool]:
    cur.execute(
        """
        SELECT id FROM games
        WHERE source = %s AND external_id = %s
        """,
        (record.source, record.external_id),
    )
    existing = cur.fetchone()
    cur.execute(
        """
        INSERT INTO games (
            source, external_id, sport, home_team, away_team, commence_time
        ) VALUES (%s, %s, %s, %s, %s, %s)
        ON CONFLICT (source, external_id) DO UPDATE
        SET
            sport = EXCLUDED.sport,
            home_team = EXCLUDED.home_team,
            away_team = EXCLUDED.away_team,
            commence_time = EXCLUDED.commence_time,
            updated_at = NOW()
        RETURNING id
        """,
        (
            record.source,
            record.external_id,
            record.sport,
            record.home_team,
            record.away_team,
            record.commence_time,
        ),
    )
    game_id = int(cur.fetchone()[0])
    return game_id, existing is None


def insert_snapshot_if_changed(cur: psycopg.Cursor, game_id: int, record: GameOddsRecord, market: MarketRow) -> tuple[bool, bool]:
    implied_probability = american_to_implied_probability(market.price_american)
    snapshot_hash = compute_snapshot_hash(record, market, implied_probability)

    cur.execute(
        """
        SELECT snapshot_hash
        FROM odds_history
        WHERE game_id = %s
          AND source = %s
          AND book_key = %s
          AND market_key = %s
          AND outcome_name = %s
          AND outcome_side = %s
          AND point IS NOT DISTINCT FROM %s
        ORDER BY captured_at DESC
        LIMIT 1
        """,
        (
            game_id,
            record.source,
            record.book_key,
            market.market_key,
            market.outcome_name,
            market.outcome_side,
            market.point,
        ),
    )
    latest = cur.fetchone()
    if latest is not None and str(latest[0]) == snapshot_hash:
        return False, True

    cur.execute(
        """
        INSERT INTO odds_history (
            game_id,
            source,
            book_key,
            book_name,
            market_key,
            market_name,
            outcome_name,
            outcome_side,
            price_american,
            point,
            implied_probability,
            snapshot_hash,
            raw_json,
            captured_at
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s
        )
        """,
        (
            game_id,
            record.source,
            record.book_key,
            record.book_name,
            market.market_key,
            market.market_name,
            market.outcome_name,
            market.outcome_side,
            market.price_american,
            market.point,
            implied_probability,
            snapshot_hash,
            json.dumps(record.raw_payload, separators=(",", ":"), ensure_ascii=False),
            record.commence_time,
        ),
    )
    return True, False


def parse_game_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value or "").strip()
    if not raw:
        raise ValueError("date is required")
    return datetime.strptime(raw, "%Y-%m-%d").date()


def parse_commence_time(game_date: date, raw_time: str) -> datetime:
    normalized = re.sub(r"\s+", " ", raw_time.strip().upper())
    try:
        dt_local = datetime.strptime(f"{game_date.isoformat()} {normalized}", "%Y-%m-%d %I:%M %p")
    except ValueError:
        dt_local = datetime(game_date.year, game_date.month, game_date.day, 19, 5)
    return dt_local.replace(tzinfo=ET_ZONE).astimezone(timezone.utc)


def parse_american(value: Any) -> Optional[int]:
    if value is None:
        return None
    raw = str(value).strip().replace("−", "-")
    if raw == "":
        return None
    if raw.startswith("+"):
        raw = raw[1:]
    try:
        parsed = int(float(raw))
    except ValueError:
        return None
    if parsed == 0:
        return None
    return parsed


def parse_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    raw = str(value).strip()
    if raw == "":
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def american_to_implied_probability(price_american: int) -> float:
    if price_american > 0:
        return 100.0 / (price_american + 100.0)
    return abs(price_american) / (abs(price_american) + 100.0)


def build_external_id(game_date: date, away_team: str, home_team: str) -> str:
    return f"mlb-{game_date.isoformat()}-{slugify(away_team)}-{slugify(home_team)}"


def compute_snapshot_hash(record: GameOddsRecord, market: MarketRow, implied: float) -> str:
    payload = {
        "external_id": record.external_id,
        "book_key": record.book_key,
        "market_key": market.market_key,
        "outcome_name": market.outcome_name,
        "outcome_side": market.outcome_side,
        "price_american": market.price_american,
        "point": market.point,
        "implied_probability": round(implied, 8),
        "upstream_source": record.upstream_source,
    }
    raw = json.dumps(payload, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def normalize_team(team_name: str) -> str:
    return re.sub(r"\s+", " ", team_name).strip()


def slugify(value: str) -> str:
    lowered = value.strip().lower()
    return re.sub(r"[^a-z0-9]+", "-", lowered).strip("-")


def source_priority(source: str) -> int:
    return SOURCE_PRIORITY.get(source.lower(), 100)


def market_precedence_key(record: GameOddsRecord, market: MarketRow) -> Tuple[str, str, str, str, Optional[float]]:
    return (record.external_id, record.book_key, market.market_key, market.outcome_side, market.point)


if __name__ == "__main__":
    main()
